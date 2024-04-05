from datetime import time
from random import randint
from openpyxl import Workbook

z = int(input("How many lines do you want?"))


wb= Workbook()
ws= wb.active

def generate():
    t = 1
    while t <= z:

        r= randint(8,9)
        n =randint(0,3)

        list_times_start = [0,15,30,45]
        p = list_times_start[n]
        a = time(r,p)

        list_times_end = [30,45,00,15]
        f = list_times_end[n]

        if n < 2:
            b= time(r+8,f)
        else:
            b= time(r+9,f)


        ws[f"A{t}"]= a
        ws[f"B{t}"] = b
        ws[f"C{t}"] = time(0,30)

        t+=1

generate()

wb.save("excel_to_copy_from.xlsx")



